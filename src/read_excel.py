from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from warnings import filterwarnings
from datetime import datetime


def get_worksheets(filename: str) -> list[Worksheet]:
    """"Get list of worksheets objects from filename.

    Args: filename(str)
    Returns: list of worksheets

    Will filter out worksheets that are not 4 characters starting with a
    digit. This means only worksheets with results data will be retained
    from Cambridge International component marks sheet.
    """

    filterwarnings(
        "ignore",
        message="Cannot parse header or footer so it will be ignored"
    )
    wb = load_workbook(filename=filename)
    sheets = []
    for sheet in wb.worksheets:
        if len(sheet.title) == 4 and sheet.title[0].isdigit():
            sheets.append(sheet)
    return sheets


def get_date(sheets: list) -> str:
    """Finds date from Cambridge International component results file

    Args: list of worksheets
    Returns date as a string formatted as yyyy-mm-dd

    If date not found returns today's date.
    """

    parsed_date = datetime.now()
    if sheets:
        sheet = sheets[0]
        for cell in sheet['A']:
            val = cell.value
            if isinstance(val, str) and "Report generated" in val:
                date = cell.value[20:]
                break
        parsed_date = datetime.strptime(date, r"%d%b%Y")
    formatted_date = datetime.strftime(parsed_date, r"%Y-%m-%d")
    return formatted_date


def get_titles(option: str, row: tuple) -> dict:
    """Find column numbers for data.

    Args:
        option(str): the option code for the exam
        row(tuple): a row of the worksheet as a tuple of Cell objects
    Returns:
        Dictionary with column number information:
            Under the components key is a list of the column number
            and the component code for the final marks for each component.
        {
            "CandidateNumber": <column number> (int)
            "Components": [
                (<column number>(int), "<option>/<component>"(str)),
                (<column number>(int), "<option>/<component>"(str))
                ...
            ]
        }
    """
    titles = {
        "Components": []
    }
    option += "/"
    for cell in row:
        if cell.value == "Candidate number":
            titles["CandidateNumber"] = cell.column
        if isinstance(cell.value, str) and "component" in cell.value.lower():
            component = option + cell.value[-2:]
            titles["Components"].append((cell.column, component))
    return titles


def get_sheet_results(sheet: Worksheet) -> list:
    """Get results data for one sheet."""
    results = []
    option = sheet.title

    # get titles
    start = None
    for row in sheet:
        if row[0].value == "Syllabus":
            start = row[0].row
            break
    if not start:
        return results
    titles = get_titles(option, sheet[start])

    # get results
    candidate_col = titles["CandidateNumber"]
    result_cols = titles["Components"]
    start += 1
    for row in range(start, 2000):
        if not sheet.cell(row, 1).value:
            break
        candidate = {}
        candidate["CandidateNumber"] = sheet.cell(row, candidate_col).value
        candidate["Results"] = []
        for col, component in result_cols:
            mark = sheet.cell(row, col).value
            candidate["Results"].append((component, mark))

        results.append(candidate)

    return results


def get_results(sheets: list) -> list:
    """Get results data."""
    results = []
    for sheet in sheets:
        results += get_sheet_results(sheet)

    return results


def get_centre_number(sheets: list) -> int:
    "Gets candidate number"
    centre_number = 0
    if sheets:
        sheet = sheets[0]
        for cell in sheet['C']:
            val = cell.value
            if isinstance(val, str) and val.lower() == "centre number":
                centre_number = sheet.cell(cell.row + 1, cell.column).value
                break
    return centre_number
